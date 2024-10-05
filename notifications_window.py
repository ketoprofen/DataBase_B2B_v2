# notifications_window.py
from PyQt5.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QLineEdit, QPushButton, QLabel, QTableView, QHeaderView, QMessageBox
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QColor, QStandardItemModel, QStandardItem, QPalette
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font

class NotificationsWindow(QWidget):
    def __init__(self, conn):
        super().__init__()
        self.setWindowTitle('Notifiche - Controlla queste targhe!!!')
        self.resize(1000, 600)
        self.conn = conn
        self.cursor = self.conn.cursor()
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()

        button_layout = QHBoxLayout()

        # Search Box
        self.search_box = QLineEdit(self)
        self.search_box.setPlaceholderText("Search by Targa...")
        self.search_box.setFixedWidth(150)
        self.search_box.textChanged.connect(self.load_notifications)  # Update notifications when text changes
        button_layout.addWidget(self.search_box)

        # Export Button
        self.export_button = QPushButton("Esporta Excel", self)
        self.export_button.clicked.connect(self.export_to_excel)
        button_layout.addWidget(self.export_button)

        button_layout.addStretch()

        layout.addLayout(button_layout)

        self.table = QTableView()
        layout.addWidget(self.table)

        # Legend
        legend_layout = QHBoxLayout()
        legend_label = QLabel("Legenda:")
        self.yellow_label = QLabel("10-15 giorni")  # Assign self
        self.yellow_label.setAutoFillBackground(True)
        yellow_palette = self.yellow_label.palette()
        yellow_palette.setColor(QPalette.Window, QColor('yellow'))
        self.yellow_label.setPalette(yellow_palette)
        self.yellow_label.setFixedSize(100, 20)
        self.yellow_label.setAlignment(Qt.AlignCenter)

        self.orange_label = QLabel("16-20 giorni")  # Assign self
        self.orange_label.setAutoFillBackground(True)
        orange_palette = self.orange_label.palette()
        orange_palette.setColor(QPalette.Window, QColor('orange'))
        self.orange_label.setPalette(orange_palette)
        self.orange_label.setFixedSize(100, 20)
        self.orange_label.setAlignment(Qt.AlignCenter)

        self.red_label = QLabel("Oltre 20 giorni")  # Assign self
        self.red_label.setAutoFillBackground(True)
        red_palette = self.red_label.palette()
        red_palette.setColor(QPalette.Window, QColor('red'))
        self.red_label.setPalette(red_palette)
        self.red_label.setFixedSize(100, 20)
        self.red_label.setAlignment(Qt.AlignCenter)

        legend_layout.addWidget(legend_label)
        legend_layout.addWidget(self.yellow_label)
        legend_layout.addWidget(self.orange_label)
        legend_layout.addWidget(self.red_label)
        legend_layout.addStretch()

        layout.addLayout(legend_layout)

        self.setLayout(layout)
        self.load_notifications()

    def load_notifications(self):
        # Get search filter from search box
        search_text = self.search_box.text().upper().strip()

        # Construct the query to filter by flotta, targa, or ditta using a single input field
        if search_text:
            query = '''
                SELECT *
                FROM records
                WHERE (flotta LIKE ? OR targa LIKE ? OR ditta LIKE ?)
                AND stato != "Consegnata"
            '''
            # Use the same search term for all three fields
            search_param = f'%{search_text}%'
            self.cursor.execute(query, (search_param, search_param, search_param))
        else:
            query = '''
                SELECT *
                FROM records
                WHERE stato != "Consegnata"
            '''
            self.cursor.execute(query)

        records = self.cursor.fetchall()

        notifications = []
        for record in records:
            data_incarico_str = record['data_incarico']
            try:
                data_incarico = datetime.strptime(data_incarico_str, '%d/%m/%Y')
                working_days = self.calculate_working_days(data_incarico, datetime.now())
                if working_days > 10:
                    record_dict = {}
                    for key in record.keys():
                        record_dict[key] = record[key]
                    record_dict['working_days'] = working_days
                    notifications.append(record_dict)
            except Exception:
                continue  # Skip records with invalid dates

        # Initialize counts for legend labels to zero
        count_10_15 = 0
        count_16_20 = 0
        count_over_20 = 0

        if notifications:
            self.df = pd.DataFrame(notifications)

            def get_color(working_days):
                if 10 < working_days <= 15:
                    return 'yellow'
                elif 15 < working_days <= 20:
                    return 'orange'
                elif working_days > 20:
                    return 'red'
                else:
                    return ''

            self.df['Color'] = self.df['working_days'].apply(get_color)

            # Calculate row counts for each interval
            count_10_15 = len(self.df[(self.df['working_days'] > 10) & (self.df['working_days'] <= 15)])
            count_16_20 = len(self.df[(self.df['working_days'] > 15) & (self.df['working_days'] <= 20)])
            count_over_20 = len(self.df[self.df['working_days'] > 20])

            standard_model = QStandardItemModel()
            headers = [col for col in self.df.columns if col != 'Color' and col.lower() != 'id']
            standard_model.setHorizontalHeaderLabels(headers)

            for _, row in self.df.iterrows():
                items = []
                for field in headers:
                    item = QStandardItem(str(row[field]))
                    items.append(item)
                standard_model.appendRow(items)

            self.table.setModel(standard_model)
            self.table.setEditTriggers(QTableView.NoEditTriggers)
            self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

            # Apply color to rows
            for row in range(standard_model.rowCount()):
                color = self.df.iloc[row]['Color']
                if color:
                    for col in range(standard_model.columnCount()):
                        item = standard_model.item(row, col)
                        if item:
                            item.setBackground(QColor(color))
        else:
            # Clear table if no records are found
            standard_model = QStandardItemModel()
            self.table.setModel(standard_model)

        # Update legend labels with counts (even if they are 0)
        self.yellow_label.setText(f"10-15 giorni ({count_10_15})")
        self.orange_label.setText(f"16-20 giorni ({count_16_20})")
        self.red_label.setText(f"Oltre 20 giorni ({count_over_20})")

    def calculate_working_days(self, start_date, end_date):
        if start_date > end_date:
            return 0
        day_generator = (start_date + timedelta(days=x) for x in range((end_date - start_date).days + 1))
        working_days = sum(1 for day in day_generator if day.weekday() < 5)
        return working_days

    def export_to_excel(self):
        if hasattr(self, 'df') and not self.df.empty:
            # Ensure 'Id' column is removed before exporting
            export_df = self.df.copy()
            export_df = export_df.loc[:, ~export_df.columns.str.lower().isin(['id', 'color'])]
            # Reorder columns to place 'data_consegnata' right after 'stato'
            columns = list(export_df.columns)
            if 'stato' in columns and 'data_consegnata' in columns:
                stato_index = columns.index('stato')
                data_consegnata_index = columns.index('data_consegnata')
                if data_consegnata_index != stato_index + 1:
                    columns.insert(stato_index + 1, columns.pop(data_consegnata_index))
                export_df = export_df[columns]

            # Create Excel workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Lavorazioni Critiche"

            # Define border style and font style
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            bold_font = Font(bold=True)

            # Write headers with bold font and borders
            for col_num, header in enumerate(export_df.columns, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.border = thin_border
                cell.font = bold_font

            # Write data with colors and borders
            for row_num, row in enumerate(export_df.itertuples(index=False), 2):
                for col_num, value in enumerate(row, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.border = thin_border
                    if hasattr(row, 'working_days'):
                        if row.working_days > 10 and row.working_days <= 15:
                            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
                        elif row.working_days > 15 and row.working_days <= 20:
                            cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange
                        elif row.working_days > 20:
                            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red

            # Create filename based on search text and current date
            search_text = self.search_box.text().strip().replace(' ', '_')
            current_date = datetime.now().strftime('%Y%m%d')
            filename = f'Lavorazioni_Critiche_{search_text}_{current_date}.xlsx' if search_text else f'Lavorazioni_critiche_all_{current_date}.xlsx'
            wb.save(filename)

            # Show message box
            QMessageBox.information(self, "Esportazione Completata Con successo!", f"File '{filename}' creato correttamente.")
            print(f"Excel file '{filename}' has been created.")